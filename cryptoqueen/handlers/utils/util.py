import io
import csv
from openpyxl import Workbook

from datetime import datetime
from datetime import timedelta
from django.db.models import QuerySet
from typing import Dict


def _get_csv_from_qs_values(queryset: QuerySet[Dict], filename: str = 'users'):
    keys = queryset[0].keys()

    # csv module can write data in io.StringIO buffer only
    s = io.StringIO()
    dict_writer = csv.DictWriter(s, fieldnames=keys)
    dict_writer.writeheader()
    dict_writer.writerows(queryset)
    s.seek(0)

    # python-telegram-bot library can send files only from io.BytesIO buffer
    # we need to convert StringIO to BytesIO
    buf = io.BytesIO()

    # extract csv-string, convert it to bytes and write to buffer
    buf.write(s.getvalue().encode())
    buf.seek(0)

    # set a filename with file's extension
    buf.name = f"{filename}__{datetime.now().strftime('%Y.%m.%d.%H.%M')}.csv"

    return buf

def _get_xlsx_from_qs_values(queryset: QuerySet[Dict], filename: str = 'users'): # need update TODO
    """
    Downloads all movies as Excel file with a single worksheet
    """
    movie_queryset = dict()
    
    # response = HttpResponse(
    #     content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    # )
    # response['Content-Disposition'] = 'attachment; filename={date}-movies.xlsx'.format(
    #     date=datetime.now().strftime('%Y-%m-%d'),
    # )
    workbook = Workbook()
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Movies'

    # Define the titles for columns
    columns = [
        'ID',
        'Title',
        'Description',
        'Length',
        'Rating',
        'Price',
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for movie in movie_queryset:
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            movie.pk,
            movie.title,
            movie.description,
            movie.length_in_minutes,
            movie.rating,
            movie.price,
        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(movie_queryset)